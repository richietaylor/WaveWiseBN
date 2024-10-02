import openpyxl
path = "data.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

row = sheet_obj.max_row - 2
column = sheet_obj.max_column - 2
counter = [0,0,0,0]

wb = openpyxl.load_workbook("out.xlsx")
sheet = wb.active


#for colI in range(2, column+1):
#    print(sheet_obj.cell(row=1, column=colI).value)
#    sheet.cell(1, colI).value = sheet_obj.cell(row=1, column=colI).value
#    counter = [0,0,0,0,0,0,0,0]
#    for rowI in range (2, row+1):
#        max = (float)(sheet_obj.cell(row=row+1, column=colI).value)
#        min = (float)(sheet_obj.cell(row=row+2, column=colI).value)
#        mid = (max+min)/2.0
#        midUp = (mid+max)/2.0
#        midDown = (min+mid)/2.0
#        currCell = (sheet_obj.cell(row=rowI, column=colI).value)
 #       if(currCell == ''): continue
#        else: currCell = (float)(currCell)
#        if(currCell < midDown): counter[0]+=1
#        elif(currCell < mid): counter[1]+=1
#        elif(currCell < midUp): counter[2]+=1
#        elif(currCell < max): counter[3]+=1
#    print(counter)
#    sheet.cell(2, colI).value = counter[0]
#    sheet.cell(3, colI).value = counter[1]
#    sheet.cell(4, colI).value = counter[2]
#    sheet.cell(5, colI).value = counter[3]
#    print()
#wb.save("out.xlsx")

for colI in range(2, column+1):
    heading = sheet_obj.cell(row=1, column=colI).value
    counter = [0,0,0,0,0,0,0,0]
    if("Direction" in heading):
        print(heading)
        for rowI in range (2, row+1):
            currCell = (sheet_obj.cell(row=rowI, column=colI).value)
            if(currCell == ''): continue
            else: currCell = (float)(currCell)
            if(currCell > 337.5 or currCell <= 22.5): counter[0]+=1
            elif(22.5 < currCell <= 67.5): counter[1]+=1
            elif(67.5 < currCell <= 112.5): counter[2]+=1
            elif(112.5 < currCell <= 157.5): counter[3]+=1
            elif(157.5 < currCell <= 202.5): counter[4]+=1
            elif(202.5 < currCell <= 247.5): counter[5]+=1
            elif(247.5 < currCell <= 292.5): counter[6]+=1
            elif(292.5 < currCell <= 337.5): counter[7]+=1
                
        print(counter)
        print()
wb.save("out.xlsx")
